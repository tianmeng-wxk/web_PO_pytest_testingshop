#from PIL import Image
import requests,logging,time
import yaml
from selenium import webdriver
from Chaojiying_Python.chaojiying import Chaojiying_Client
#第三方接口识别验证码方法一
def ivercode(driver):
    driver.save_screenshot("E:\\pycharm\\testingshop\\vercode_img\\page.png")
    vcode = driver.find_element_by_xpath("//*[@id='verify_code_img']")
    loc = vcode.location
    size = vcode.size
    left = loc['x']
    top = loc['y']
    right = (loc['x'] + size['width'])
    button = (loc['y'] + size['height'])
    page_pic = Image.open("E:\\pycharm\\testingshop\\vercode_img\\page.png")
    v_code_pic = page_pic.crop((left, top, right, button))
    v_code_pic.save("E:\\pycharm\\testingshop\\vercode_img\\vercode.png")
    chaojiying = Chaojiying_Client('qq121292679', 'a546245426', '904603')  # 用户中心>>软件ID 生成一个替换 96001
    im = open('E:\\pycharm\\testingshop\\vercode_img\\vercode.png', 'rb').read()  # 本地图片文件路径 来替换 a.jpg 有时WIN系统须要//
    res = chaojiying.PostPic(im, 1902)
    vercode = res['pic_str']
    return vercode
#第三方接口识别验证码方法二
def ivercode2(driver):
    ele = driver.find_element_by_xpath("//*[@id='verify_code_img']")
    ele.screenshot("E:\\pycharm\\testingshop\\vercode_img\\verify.png")
    headers = {
        'Connection': 'Keep-Alive',
        'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0)',
    }
    data = {
        'user': 'wuqingfqng',
        'pass2': '6e8ebd2e301f3d5331e1e230ff3f3ca5',#密碼：wuqing&fqng
        "softid": "904357",
        "codetype": "1902"
    }
    userfile = open("E:\\pycharm\\testingshop\\vercode_img\\verify.png","rb").read()
    userfile = {"userfile": ("E:\\pycharm\\testingshop\\vercode_img\\verify.png", userfile)}
    res = requests.post("http://upload.chaojiying.net/Upload/Processing.php",data=data,files=userfile,headers=headers)
    res = res.json()
    vercode = res["pic_str"]
    return vercode


def read_yaml(file_path):
    file = open(file_path, encoding='utf-8')
    testdata = yaml.load(file, Loader=yaml.FullLoader)  # 或者yaml.full_load()
    return testdata


class Logger:
    def log(self):
        logger=logging.getLogger("logger")
        logger.setLevel(logging.DEBUG)
        if not logger.handlers:
            sh=logging.StreamHandler()
            fh=logging.FileHandler(filename="../config/log/{}_log".format(time.strftime("%Y_%m_%d_%H_%M_%S",time.localtime())),encoding="utf-8")
            formator=logging.Formatter(fmt="%(asctime)s %(filename)s %(levelname)s %(msg)s", datefmt="%Y/%m%d%X")
            sh.setFormatter(formator)
            fh.setFormatter(formator)
            logger.addHandler(fh)
            logger.addHandler(sh)
        return logger

def browser_type(type):
    type=type.upper()
    if type == "CHR":
        driver = webdriver.Chrome()
    elif type == "IE":
        driver = webdriver.Ie()
    elif type == "FF":
        driver = webdriver.Firefox()
    return driver

#xlrd读取excel方法1
import xlrd
#import xlutils
class ReadExcel:
    def __init__(self, excel_path, sheet_name):
        self.workbook = xlrd.open_workbook(excel_path)
        self.worksheet = self.workbook.sheet_by_name(sheet_name)
        self.rownum = self.worksheet.nrows
        self.colnum = self.worksheet.ncols

    def dict_data(self):
        if self.rownum <= 1:
             print("表格行数小于等于1，不能进行自动化")
        else:
             list = []
             self.headers = self.worksheet.row_values(0)
             #self.headers = self.worksheet.row_values(0)
             #j = 1#从1开始
             for i in range(1, self.rownum):
                 s = {}
                 values = self.worksheet.row_values(i)

                 for x in range(self.colnum):
                    s[self.headers[x]] = values[x]
                 list.append(s)
                 #j += 1
             return list


#xlrd读取excel方法2
def get_data(wb_path,ws_name):
    workbook = xlrd.open_workbook(wb_path)
    worksheet = workbook.sheet_by_name(ws_name)
    sheetdata=[]
    headers = worksheet.row_values(0)
    for r in range(1, worksheet.nrows):#或者worksheet.nrows-1
        row_vars = worksheet.row_values(r)
        tmp_dict = {}
        # tmp_dict[headers[0]] = row_vars[0]
        # tmp_dict[headers[1]] = row_vars[1]
        # tmp_dict[headers[2]] = row_vars[2]
        # tmp_dict[headers[3]] = row_vars[3]
        # for i in range(len(headers)):
        #     tmp_dict[headers[i]] = row_vars[i]
        tmp_dict = dict(zip(headers, row_vars))
        sheetdata.append(tmp_dict)
    return sheetdata


import openpyxl
#openpyxl加载excel
def load_excel(excel_path):
    global excel
    excel = openpyxl.load_workbook(excel_path)
    return excel
#openpyxl记载sheet
def load_sheet(sheet_name):
    sheet = excel[sheet_name]
    return sheet
#openpyxl读取数据
def read_excel(excel_path, sheet_name):
    excel = load_excel(excel_path)
    sheet = load_sheet(sheet_name)
    l = []
    for i in range(2, sheet.max_row+1):
        data = []
        arg1 = sheet.cell(i, 1).value
        arg2 = sheet.cell(i, 2).value
        data.append(arg1)
        data.append(arg2)
        l.append(data)
    return l
#read_excel("../config/order_check.xlsx", "Sheet1")



